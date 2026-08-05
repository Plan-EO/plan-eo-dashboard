"""
Syncs text metadata fields from Plan-EO_dashboard_raster_metadata.xlsx into
raster-data/raster-layers.json and app/src/lib/data/raster-layers-fallback.json.

Matching key: Excel `File_name` == stem of JSON layer `path`  (e.g. Flr_Fin_Pr)

Fields synced (Excel column → JSON field):
  Panel_heading    → panelHeading
  Panel_subheading → panelSubheading
  Popup_Heading    → popupHeading
  Popup_subheading → popupSubheading
  Indicator        → indicator
  Duration         → period
  Definition       → definition
  Source           → study
  Hyperlink        → hyperlink
"""

import json
import os
import openpyxl

EXCEL_FILE = "Plan-EO_dashboard_raster_metadata.xlsx"
JSON_FILES = [
    "raster-data/raster-layers.json",
    "app/src/lib/data/raster-layers-fallback.json",
]

COLUMN_MAP = {
    "Panel_heading":    "panelHeading",
    "Panel_subheading": "panelSubheading",
    "Layer_description": "layerDescription",
    "Popup_Heading":    "popupHeading",
    "Popup_subheading": "popupSubheading",
    "Indicator":        "indicator",
    "Duration":         "period",
    "Definition":       "definition",
    "Source":           "study",
    "Hyperlink":        "hyperlink",
    "Footnote":         "footnoteDetail",
    "Layer_Manager":    "layerManagerCategory",
}


def load_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        file_name = record.get("File_name")
        if file_name:
            rows[file_name] = record
    return rows


def stem(path_str):
    """Return the filename without extension from a path like '02_Risk_factors/Floor/Flr_Fin_Pr.tif'."""
    return os.path.splitext(os.path.basename(path_str))[0]


def sync_json(json_path, excel_rows):
    with open(json_path) as f:
        data = json.load(f)

    updated = 0
    unmatched = []

    for layer in data["layers"]:
        key = stem(layer.get("path", ""))
        if key not in excel_rows:
            unmatched.append(key)
            continue

        record = excel_rows[key]
        for excel_col, json_field in COLUMN_MAP.items():
            value = record.get(excel_col)
            if value is not None:
                layer[json_field] = str(value).strip()
            elif json_field in layer:
                del layer[json_field]
        updated += 1

    with open(json_path, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"{json_path}: {updated} layers updated, {len(unmatched)} unmatched")
    if unmatched:
        print(f"  Unmatched stems: {', '.join(unmatched)}")


if __name__ == "__main__":
    excel_rows = load_excel(EXCEL_FILE)
    print(f"Loaded {len(excel_rows)} rows from {EXCEL_FILE}\n")
    for json_file in JSON_FILES:
        sync_json(json_file, excel_rows)
